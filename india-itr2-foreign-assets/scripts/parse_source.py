#!/usr/bin/env python3
"""Parse common tax-source formats into one normalized JSON envelope.

The parser performs deterministic structural extraction only. It does not
classify tax items or create reconciled facts.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import html
import importlib.util
import io
import json
import os
import shutil
import subprocess
import sys
import tempfile
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, timezone
from html.parser import HTMLParser
from pathlib import Path, PurePosixPath
from typing import Any
from xml.etree import ElementTree as ET


SCHEMA_VERSION = "1.0"
PARSER_NAME = "itr-generic-source-parser"
PARSER_VERSION = "1.0.0"


class ParseError(Exception):
    pass


@dataclass
class Limits:
    max_file_bytes: int = 512 * 1024 * 1024
    max_unit_text_chars: int = 5_000_000
    max_rows: int = 250_000
    max_json_leaves: int = 500_000
    max_archive_members: int = 2_000
    max_archive_uncompressed: int = 2 * 1024 * 1024 * 1024
    max_member_bytes: int = 512 * 1024 * 1024
    max_archive_depth: int = 2


@dataclass
class Options:
    password: str | None = None
    ocr: str = "auto"
    limits: Limits | None = None


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def atomic_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fd, temporary = tempfile.mkstemp(
        dir=path.parent, prefix=f".{path.name}.", suffix=".tmp"
    )
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as handle:
            json.dump(payload, handle, indent=2, sort_keys=True, ensure_ascii=False)
            handle.write("\n")
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary, path)
    finally:
        if os.path.exists(temporary):
            os.unlink(temporary)


def json_safe(value: Any) -> Any:
    if value is None or isinstance(value, (bool, int, float, str)):
        return value
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    return str(value)


def decode_text(data: bytes) -> tuple[str, str]:
    if data.startswith(b"\xef\xbb\xbf"):
        return data.decode("utf-8-sig"), "utf-8-sig"
    if data.startswith((b"\xff\xfe", b"\xfe\xff")):
        return data.decode("utf-16"), "utf-16"
    for encoding in ("utf-8", "cp1252", "latin-1"):
        try:
            return data.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace"), "utf-8-replacement"


def local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def detect_format(data: bytes, name: str) -> str:
    extension = Path(name).suffix.lower()
    extension_map = {
        ".pdf": "pdf",
        ".json": "json",
        ".csv": "csv",
        ".tsv": "tsv",
        ".txt": "text",
        ".log": "text",
        ".md": "text",
        ".html": "html",
        ".htm": "html",
        ".xml": "xml",
        ".xlsx": "xlsx",
        ".xlsm": "xlsx",
        ".xltx": "xlsx",
        ".xls": "xls",
        ".docx": "docx",
        ".zip": "zip",
        ".png": "image",
        ".jpg": "image",
        ".jpeg": "image",
        ".tif": "image",
        ".tiff": "image",
        ".bmp": "image",
        ".webp": "image",
    }
    if extension in extension_map:
        return extension_map[extension]
    if data.startswith(b"%PDF-"):
        return "pdf"
    if data.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"):
        return "xls"
    if data.startswith((b"PK\x03\x04", b"PK\x05\x06", b"PK\x07\x08")):
        try:
            with zipfile.ZipFile(io.BytesIO(data)) as archive:
                names = set(archive.namelist())
            if "xl/workbook.xml" in names:
                return "xlsx"
            if "word/document.xml" in names:
                return "docx"
        except zipfile.BadZipFile:
            pass
        return "zip"
    stripped = data.lstrip()
    if stripped.startswith((b"{", b"[")):
        try:
            json.loads(decode_text(data)[0])
            return "json"
        except (json.JSONDecodeError, UnicodeDecodeError):
            pass
    if stripped.startswith(b"<"):
        head = stripped[:500].lower()
        return "html" if b"<html" in head or b"<!doctype html" in head else "xml"
    return "text"


def truncate_text(text: str, limit: int, warnings: list[str], label: str) -> str:
    if len(text) <= limit:
        return text
    warnings.append(f"{label} text truncated from {len(text)} to {limit} characters")
    return text[:limit]


def parse_json(data: bytes, options: Options) -> dict[str, Any]:
    text, encoding = decode_text(data)
    try:
        payload = json.loads(text)
    except json.JSONDecodeError as exc:
        raise ParseError(f"invalid JSON: {exc}") from exc
    limits = options.limits or Limits()
    warnings: list[str] = []
    leaves: list[dict[str, Any]] = []
    stack: list[tuple[str, Any]] = [("", payload)]
    while stack:
        pointer, value = stack.pop()
        if isinstance(value, dict):
            for key, child in reversed(list(value.items())):
                escaped = str(key).replace("~", "~0").replace("/", "~1")
                stack.append((f"{pointer}/{escaped}", child))
        elif isinstance(value, list):
            for index in range(len(value) - 1, -1, -1):
                stack.append((f"{pointer}/{index}", value[index]))
        else:
            leaves.append(
                {
                    "json_pointer": pointer or "/",
                    "value": json_safe(value),
                    "value_type": type(value).__name__,
                }
            )
            if len(leaves) >= limits.max_json_leaves:
                warnings.append(
                    f"JSON leaf limit reached at {limits.max_json_leaves}; "
                    "rerun with a larger limit"
                )
                break
    chunks = [
        {
            "unit_id": f"json-leaves-{start // 5000 + 1}",
            "kind": "json_leaves",
            "locator": {"json_pointer": "/"},
            "records": leaves[start : start + 5000],
        }
        for start in range(0, len(leaves), 5000)
    ]
    return {
        "format": "json",
        "backend": "python-json",
        "metadata": {
            "encoding": encoding,
            "root_type": type(payload).__name__,
            "leaf_count": len(leaves),
        },
        "units": chunks,
        "members": [],
        "warnings": warnings,
    }


def parse_delimited(data: bytes, options: Options, delimiter: str | None) -> dict[str, Any]:
    text, encoding = decode_text(data)
    limits = options.limits or Limits()
    warnings: list[str] = []
    sample = text[:65536]
    if delimiter is None:
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = ","
            warnings.append("Could not detect delimiter; defaulted to comma")
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    rows: list[dict[str, Any]] = []
    for row_number, row in enumerate(reader, start=1):
        if row_number > limits.max_rows:
            warnings.append(
                f"Row limit reached at {limits.max_rows}; rerun with a larger limit"
            )
            break
        rows.append(
            {
                "row": row_number,
                "cells": [
                    {
                        "column": index + 1,
                        "column_name": spreadsheet_column(index + 1),
                        "value": value,
                    }
                    for index, value in enumerate(row)
                ],
            }
        )
    return {
        "format": "delimited",
        "backend": "python-csv",
        "metadata": {
            "encoding": encoding,
            "delimiter": delimiter,
            "row_count": len(rows),
        },
        "units": [
            {
                "unit_id": "table-1",
                "kind": "table",
                "locator": {"table": 1},
                "rows": rows,
            }
        ],
        "members": [],
        "warnings": warnings,
    }


def spreadsheet_column(number: int) -> str:
    output = ""
    while number:
        number, remainder = divmod(number - 1, 26)
        output = chr(65 + remainder) + output
    return output


def parse_xlsx_openpyxl(data: bytes, options: Options) -> dict[str, Any]:
    try:
        import openpyxl
    except ImportError as exc:
        raise ParseError("openpyxl is unavailable") from exc
    limits = options.limits or Limits()
    warnings: list[str] = []
    try:
        workbook = openpyxl.load_workbook(
            io.BytesIO(data), read_only=True, data_only=False
        )
    except Exception as exc:
        raise ParseError(f"openpyxl could not open workbook: {exc}") from exc
    units: list[dict[str, Any]] = []
    total_rows = 0
    for sheet in workbook.worksheets:
        rows: list[dict[str, Any]] = []
        for row_number, row in enumerate(sheet.iter_rows(), start=1):
            if total_rows >= limits.max_rows:
                warnings.append(
                    f"Workbook row limit reached at {limits.max_rows}; "
                    "rerun with a larger limit"
                )
                break
            cells = []
            for cell in row:
                if cell.value is None:
                    continue
                cells.append(
                    {
                        "cell": cell.coordinate,
                        "column": cell.column,
                        "value": json_safe(cell.value),
                        "value_type": cell.data_type,
                        "number_format": cell.number_format,
                    }
                )
            if cells:
                rows.append({"row": row_number, "cells": cells})
            total_rows += 1
        units.append(
            {
                "unit_id": f"sheet-{len(units) + 1}",
                "kind": "sheet",
                "title": sheet.title,
                "locator": {"sheet": sheet.title},
                "rows": rows,
            }
        )
        if total_rows >= limits.max_rows:
            break
    workbook.close()
    return {
        "format": "xlsx",
        "backend": "openpyxl",
        "metadata": {"sheet_count": len(units), "row_count": total_rows},
        "units": units,
        "members": [],
        "warnings": warnings,
    }


def read_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    try:
        root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    except KeyError:
        return []
    return [
        "".join(node.text or "" for node in item.iter() if local_name(node.tag) == "t")
        for item in root
    ]


def workbook_sheet_targets(archive: zipfile.ZipFile) -> list[tuple[str, str]]:
    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    relations = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    relation_map = {
        node.attrib["Id"]: node.attrib["Target"]
        for node in relations
        if "Id" in node.attrib and "Target" in node.attrib
    }
    sheets: list[tuple[str, str]] = []
    for node in workbook.iter():
        if local_name(node.tag) != "sheet":
            continue
        relation_id = next(
            (value for key, value in node.attrib.items() if key.endswith("}id")),
            None,
        )
        if not relation_id or relation_id not in relation_map:
            continue
        target = relation_map[relation_id].lstrip("/")
        if not target.startswith("xl/"):
            target = "xl/" + target
        target = str(PurePosixPath(target))
        sheets.append((node.attrib.get("name", f"Sheet{len(sheets) + 1}"), target))
    return sheets


def xlsx_cell_value(cell: ET.Element, shared: list[str]) -> tuple[Any, str]:
    cell_type = cell.attrib.get("t", "n")
    value_node = next(
        (node for node in cell if local_name(node.tag) == "v"), None
    )
    inline_text = "".join(
        node.text or "" for node in cell.iter() if local_name(node.tag) == "t"
    )
    raw = value_node.text if value_node is not None else inline_text
    if raw is None:
        return None, cell_type
    if cell_type == "s":
        try:
            return shared[int(raw)], "shared_string"
        except (ValueError, IndexError):
            return raw, "shared_string_invalid_index"
    if cell_type in {"inlineStr", "str"}:
        return raw, "string"
    if cell_type == "b":
        return raw == "1", "boolean"
    try:
        number = float(raw)
        return int(number) if number.is_integer() else number, "number"
    except ValueError:
        return raw, cell_type


def parse_xlsx_stdlib(data: bytes, options: Options) -> dict[str, Any]:
    limits = options.limits or Limits()
    warnings = [
        "openpyxl unavailable; used OOXML fallback without style/date conversion"
    ]
    units: list[dict[str, Any]] = []
    total_rows = 0
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            shared = read_shared_strings(archive)
            for title, target in workbook_sheet_targets(archive):
                root = ET.fromstring(archive.read(target))
                rows: list[dict[str, Any]] = []
                for row_node in root.iter():
                    if local_name(row_node.tag) != "row":
                        continue
                    if total_rows >= limits.max_rows:
                        warnings.append(
                            f"Workbook row limit reached at {limits.max_rows}"
                        )
                        break
                    row_number = int(row_node.attrib.get("r", total_rows + 1))
                    cells = []
                    for cell in row_node:
                        if local_name(cell.tag) != "c":
                            continue
                        value, value_type = xlsx_cell_value(cell, shared)
                        if value is None:
                            continue
                        cells.append(
                            {
                                "cell": cell.attrib.get("r"),
                                "value": value,
                                "value_type": value_type,
                            }
                        )
                    if cells:
                        rows.append({"row": row_number, "cells": cells})
                    total_rows += 1
                units.append(
                    {
                        "unit_id": f"sheet-{len(units) + 1}",
                        "kind": "sheet",
                        "title": title,
                        "locator": {"sheet": title},
                        "rows": rows,
                    }
                )
                if total_rows >= limits.max_rows:
                    break
    except (zipfile.BadZipFile, KeyError, ET.ParseError) as exc:
        raise ParseError(f"invalid XLSX/OOXML workbook: {exc}") from exc
    return {
        "format": "xlsx",
        "backend": "python-ooxml",
        "metadata": {"sheet_count": len(units), "row_count": total_rows},
        "units": units,
        "members": [],
        "warnings": warnings,
    }


def parse_xlsx(data: bytes, options: Options) -> dict[str, Any]:
    try:
        return parse_xlsx_openpyxl(data, options)
    except ParseError as exc:
        fallback = parse_xlsx_stdlib(data, options)
        fallback["warnings"].insert(0, f"openpyxl path unavailable: {exc}")
        return fallback


def parse_xls(data: bytes, options: Options) -> dict[str, Any]:
    try:
        import xlrd
    except ImportError as exc:
        raise ParseError(
            "legacy XLS requires xlrd; install parser-requirements.txt"
        ) from exc
    limits = options.limits or Limits()
    warnings: list[str] = []
    try:
        workbook = xlrd.open_workbook(file_contents=data, on_demand=True)
    except Exception as exc:
        raise ParseError(f"xlrd could not open workbook: {exc}") from exc
    units: list[dict[str, Any]] = []
    total_rows = 0
    for sheet in workbook.sheets():
        rows = []
        for row_index in range(sheet.nrows):
            if total_rows >= limits.max_rows:
                warnings.append(f"Workbook row limit reached at {limits.max_rows}")
                break
            cells = []
            for column_index in range(sheet.ncols):
                cell = sheet.cell(row_index, column_index)
                if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                    continue
                value = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    value = xlrd.xldate_as_datetime(value, workbook.datemode).isoformat()
                elif cell.ctype == xlrd.XL_CELL_NUMBER and float(value).is_integer():
                    value = int(value)
                cells.append(
                    {
                        "cell": f"{spreadsheet_column(column_index + 1)}{row_index + 1}",
                        "column": column_index + 1,
                        "value": json_safe(value),
                        "value_type": str(cell.ctype),
                    }
                )
            if cells:
                rows.append({"row": row_index + 1, "cells": cells})
            total_rows += 1
        units.append(
            {
                "unit_id": f"sheet-{len(units) + 1}",
                "kind": "sheet",
                "title": sheet.name,
                "locator": {"sheet": sheet.name},
                "rows": rows,
            }
        )
        if total_rows >= limits.max_rows:
            break
    workbook.release_resources()
    return {
        "format": "xls",
        "backend": "xlrd",
        "metadata": {"sheet_count": len(units), "row_count": total_rows},
        "units": units,
        "members": [],
        "warnings": warnings,
    }


def ocr_image_bytes(data: bytes, suffix: str) -> tuple[str, str]:
    executable = shutil.which("tesseract")
    if not executable:
        raise ParseError("OCR requested but tesseract is unavailable")
    with tempfile.NamedTemporaryFile(suffix=suffix) as source:
        source.write(data)
        source.flush()
        result = subprocess.run(
            [executable, source.name, "stdout"],
            check=False,
            capture_output=True,
            text=True,
        )
    if result.returncode:
        raise ParseError(f"tesseract failed: {result.stderr.strip()}")
    return result.stdout, "tesseract"


def parse_pdf(data: bytes, options: Options) -> dict[str, Any]:
    warnings: list[str] = []
    limits = options.limits or Limits()
    try:
        import fitz
    except ImportError:
        fitz = None
    units: list[dict[str, Any]] = []
    metadata: dict[str, Any] = {}
    if fitz is not None:
        try:
            document = fitz.open(stream=data, filetype="pdf")
            if document.needs_pass:
                if not options.password:
                    raise ParseError(
                        "PDF is encrypted; supply a password through --password-env"
                    )
                if not document.authenticate(options.password):
                    raise ParseError("PDF password was rejected")
            metadata = {key: json_safe(value) for key, value in document.metadata.items()}
            for index, page in enumerate(document):
                text = page.get_text("text")
                blocks = [
                    {
                        "bbox": [round(value, 3) for value in block[:4]],
                        "text": block[4],
                        "block_number": block[5],
                        "block_type": block[6],
                    }
                    for block in page.get_text("blocks")
                ]
                if not text.strip() and options.ocr in {"auto", "always"}:
                    executable = shutil.which("tesseract")
                    if executable:
                        pixmap = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
                        try:
                            text, _ = ocr_image_bytes(pixmap.tobytes("png"), ".png")
                            warnings.append(f"Page {index + 1} required OCR")
                        except ParseError as exc:
                            warnings.append(f"Page {index + 1} OCR failed: {exc}")
                    else:
                        warnings.append(
                            f"Page {index + 1} has no text and tesseract is unavailable"
                        )
                text = truncate_text(
                    text,
                    limits.max_unit_text_chars,
                    warnings,
                    f"PDF page {index + 1}",
                )
                units.append(
                    {
                        "unit_id": f"page-{index + 1}",
                        "kind": "page",
                        "locator": {"page": index + 1},
                        "text": text,
                        "blocks": blocks,
                    }
                )
            document.close()
        except ParseError:
            raise
        except Exception as exc:
            raise ParseError(f"PyMuPDF could not parse PDF: {exc}") from exc
        backend = "pymupdf"
    else:
        executable = shutil.which("pdftotext")
        if not executable:
            raise ParseError(
                "PDF parsing requires PyMuPDF or pdftotext; "
                "install parser-requirements.txt"
            )
        if options.password:
            raise ParseError(
                "Encrypted PDF fallback is disabled; install PyMuPDF"
            )
        result = subprocess.run(
            [executable, "-layout", "-", "-"],
            input=data,
            check=False,
            capture_output=True,
        )
        if result.returncode:
            raise ParseError(result.stderr.decode("utf-8", errors="replace").strip())
        text, encoding = decode_text(result.stdout)
        metadata["encoding"] = encoding
        for index, page_text in enumerate(text.split("\f"), start=1):
            if not page_text and index > 1:
                continue
            units.append(
                {
                    "unit_id": f"page-{index}",
                    "kind": "page",
                    "locator": {"page": index},
                    "text": truncate_text(
                        page_text,
                        limits.max_unit_text_chars,
                        warnings,
                        f"PDF page {index}",
                    ),
                    "blocks": [],
                }
            )
        backend = "pdftotext"
    if not any(unit.get("text", "").strip() for unit in units):
        warnings.append("PDF produced no machine-readable text")
    metadata["page_count"] = len(units)
    return {
        "format": "pdf",
        "backend": backend,
        "metadata": metadata,
        "units": units,
        "members": [],
        "warnings": warnings,
    }


def parse_docx(data: bytes, options: Options) -> dict[str, Any]:
    warnings: list[str] = []
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            root = ET.fromstring(archive.read("word/document.xml"))
    except (zipfile.BadZipFile, KeyError, ET.ParseError) as exc:
        raise ParseError(f"invalid DOCX: {exc}") from exc
    units: list[dict[str, Any]] = []
    paragraph_number = 0
    table_number = 0
    for child in root.iter():
        tag = local_name(child.tag)
        if tag == "p":
            text = "".join(
                node.text or "" for node in child.iter() if local_name(node.tag) == "t"
            )
            if text.strip():
                paragraph_number += 1
                units.append(
                    {
                        "unit_id": f"paragraph-{paragraph_number}",
                        "kind": "paragraph",
                        "locator": {"paragraph": paragraph_number},
                        "text": text,
                    }
                )
        elif tag == "tbl":
            table_number += 1
            rows = []
            for row_number, row_node in enumerate(
                [node for node in child.iter() if local_name(node.tag) == "tr"],
                start=1,
            ):
                cells = []
                for column_number, cell_node in enumerate(
                    [
                        node
                        for node in row_node
                        if local_name(node.tag) == "tc"
                    ],
                    start=1,
                ):
                    text = "".join(
                        node.text or ""
                        for node in cell_node.iter()
                        if local_name(node.tag) == "t"
                    )
                    cells.append({"column": column_number, "value": text})
                rows.append({"row": row_number, "cells": cells})
            units.append(
                {
                    "unit_id": f"table-{table_number}",
                    "kind": "table",
                    "locator": {"table": table_number},
                    "rows": rows,
                }
            )
    return {
        "format": "docx",
        "backend": "python-docx-ooxml",
        "metadata": {
            "paragraph_count": paragraph_number,
            "table_count": table_number,
        },
        "units": units,
        "members": [],
        "warnings": warnings,
    }


class TextHTMLParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.parts: list[str] = []

    def handle_data(self, data: str) -> None:
        if data.strip():
            self.parts.append(html.unescape(data))


def parse_html(data: bytes, options: Options) -> dict[str, Any]:
    text, encoding = decode_text(data)
    parser = TextHTMLParser()
    parser.feed(text)
    extracted = "\n".join(parser.parts)
    warnings: list[str] = []
    extracted = truncate_text(
        extracted,
        (options.limits or Limits()).max_unit_text_chars,
        warnings,
        "HTML",
    )
    return {
        "format": "html",
        "backend": "python-html-parser",
        "metadata": {"encoding": encoding},
        "units": [
            {
                "unit_id": "document-text",
                "kind": "text",
                "locator": {"document": 1},
                "text": extracted,
            }
        ],
        "members": [],
        "warnings": warnings,
    }


def parse_xml(data: bytes, options: Options) -> dict[str, Any]:
    text, encoding = decode_text(data)
    try:
        root = ET.fromstring(text)
    except ET.ParseError as exc:
        raise ParseError(f"invalid XML: {exc}") from exc
    records: list[dict[str, Any]] = []

    def visit(node: ET.Element, path: str) -> None:
        current = f"{path}/{local_name(node.tag)}"
        if node.text and node.text.strip():
            records.append({"xml_path": current, "value": node.text.strip()})
        for key, value in node.attrib.items():
            records.append(
                {"xml_path": f"{current}/@{local_name(key)}", "value": value}
            )
        for child in node:
            visit(child, current)

    visit(root, "")
    return {
        "format": "xml",
        "backend": "python-elementtree",
        "metadata": {"encoding": encoding, "record_count": len(records)},
        "units": [
            {
                "unit_id": "xml-records-1",
                "kind": "xml_records",
                "locator": {"xml_path": f"/{local_name(root.tag)}"},
                "records": records,
            }
        ],
        "members": [],
        "warnings": [],
    }


def parse_text(data: bytes, options: Options) -> dict[str, Any]:
    if data and data.count(b"\x00") / len(data) > 0.02:
        raise ParseError(
            "unsupported binary input; add a reusable format backend"
        )
    text, encoding = decode_text(data)
    warnings: list[str] = []
    text = truncate_text(
        text,
        (options.limits or Limits()).max_unit_text_chars,
        warnings,
        "Text document",
    )
    return {
        "format": "text",
        "backend": "python-text",
        "metadata": {"encoding": encoding, "line_count": text.count("\n") + 1},
        "units": [
            {
                "unit_id": "document-text",
                "kind": "text",
                "locator": {"document": 1},
                "text": text,
            }
        ],
        "members": [],
        "warnings": warnings,
    }


def parse_image(data: bytes, name: str, options: Options) -> dict[str, Any]:
    metadata: dict[str, Any] = {}
    warnings: list[str] = []
    try:
        from PIL import Image

        with Image.open(io.BytesIO(data)) as image:
            metadata.update(
                {
                    "width": image.width,
                    "height": image.height,
                    "mode": image.mode,
                    "image_format": image.format,
                }
            )
    except ImportError:
        warnings.append("Pillow unavailable; image metadata was not extracted")
    except Exception as exc:
        warnings.append(f"Image metadata extraction failed: {exc}")
    text = ""
    if options.ocr in {"auto", "always"}:
        try:
            text, backend = ocr_image_bytes(data, Path(name).suffix or ".png")
            metadata["ocr_backend"] = backend
        except ParseError as exc:
            warnings.append(str(exc))
    return {
        "format": "image",
        "backend": metadata.get("ocr_backend", "image-metadata-only"),
        "metadata": metadata,
        "units": [
            {
                "unit_id": "image-1",
                "kind": "image",
                "locator": {"image": 1},
                "text": text,
            }
        ],
        "members": [],
        "warnings": warnings,
    }


def safe_member_name(name: str) -> bool:
    path = PurePosixPath(name)
    return not path.is_absolute() and ".." not in path.parts


def parse_zip(
    data: bytes, name: str, options: Options, depth: int
) -> dict[str, Any]:
    limits = options.limits or Limits()
    warnings: list[str] = []
    members: list[dict[str, Any]] = []
    if depth >= limits.max_archive_depth:
        raise ParseError(f"archive recursion limit reached at {depth}")
    try:
        archive = zipfile.ZipFile(io.BytesIO(data))
    except zipfile.BadZipFile as exc:
        raise ParseError(f"invalid ZIP: {exc}") from exc
    infos = archive.infolist()
    if len(infos) > limits.max_archive_members:
        warnings.append(
            f"Archive has {len(infos)} members; only the first "
            f"{limits.max_archive_members} were considered"
        )
        infos = infos[: limits.max_archive_members]
    total_uncompressed = 0
    password = options.password.encode("utf-8") if options.password else None
    aes_archive = None
    if password:
        try:
            import pyzipper

            aes_archive = pyzipper.AESZipFile(io.BytesIO(data))
            aes_archive.setpassword(password)
        except (ImportError, OSError, RuntimeError):
            aes_archive = None
    for info in infos:
        if info.is_dir():
            continue
        if not safe_member_name(info.filename):
            warnings.append(f"Skipped unsafe archive member: {info.filename}")
            continue
        total_uncompressed += info.file_size
        if total_uncompressed > limits.max_archive_uncompressed:
            warnings.append("Archive uncompressed-size limit reached")
            break
        if info.file_size > limits.max_member_bytes:
            warnings.append(f"Skipped oversized archive member: {info.filename}")
            continue
        if info.compress_size and info.file_size / info.compress_size > 1000:
            warnings.append(f"Skipped suspicious compression ratio: {info.filename}")
            continue
        try:
            member_data = archive.read(info, pwd=password)
        except (RuntimeError, NotImplementedError) as exc:
            if aes_archive is None:
                warnings.append(f"Could not read {info.filename}: {exc}")
                continue
            try:
                member_data = aes_archive.read(info.filename)
            except Exception as aes_exc:
                warnings.append(f"Could not read {info.filename}: {aes_exc}")
                continue
        try:
            member_document = parse_document(
                member_data, info.filename, options, depth=depth + 1
            )
            members.append(
                {
                    "member_path": info.filename,
                    "sha256": sha256_bytes(member_data),
                    "size": len(member_data),
                    "status": status_for_document(member_document),
                    "document": member_document,
                }
            )
        except ParseError as exc:
            members.append(
                {
                    "member_path": info.filename,
                    "sha256": sha256_bytes(member_data),
                    "size": len(member_data),
                    "status": "FAILED",
                    "error": str(exc),
                }
            )
    archive.close()
    if aes_archive is not None:
        aes_archive.close()
    return {
        "format": "zip",
        "backend": "python-zipfile",
        "metadata": {
            "member_count": len(infos),
            "parsed_member_count": len(members),
            "uncompressed_bytes_considered": total_uncompressed,
        },
        "units": [],
        "members": members,
        "warnings": warnings,
    }


def parse_document(
    data: bytes, name: str, options: Options, depth: int = 0
) -> dict[str, Any]:
    limits = options.limits or Limits()
    if len(data) > limits.max_file_bytes and depth == 0:
        raise ParseError(
            f"source exceeds max_file_bytes ({len(data)} > {limits.max_file_bytes})"
        )
    format_name = detect_format(data, name)
    if format_name == "json":
        return parse_json(data, options)
    if format_name == "csv":
        return parse_delimited(data, options, None)
    if format_name == "tsv":
        return parse_delimited(data, options, "\t")
    if format_name == "xlsx":
        return parse_xlsx(data, options)
    if format_name == "xls":
        return parse_xls(data, options)
    if format_name == "pdf":
        return parse_pdf(data, options)
    if format_name == "docx":
        return parse_docx(data, options)
    if format_name == "html":
        return parse_html(data, options)
    if format_name == "xml":
        return parse_xml(data, options)
    if format_name == "image":
        return parse_image(data, name, options)
    if format_name == "zip":
        return parse_zip(data, name, options, depth)
    return parse_text(data, options)


def status_for_document(document: dict[str, Any]) -> str:
    warnings = document.get("warnings", [])
    member_statuses = [
        member.get("status") for member in document.get("members", [])
    ]
    if member_statuses and all(status == "FAILED" for status in member_statuses):
        return "FAILED"
    if warnings or "FAILED" in member_statuses or "PARTIAL" in member_statuses:
        return "PARTIAL"
    return "COMPLETE"


def parse_path(path: Path, options: Options | None = None) -> dict[str, Any]:
    options = options or Options(limits=Limits())
    if options.limits is None:
        options.limits = Limits()
    resolved = path.expanduser().resolve()
    data = resolved.read_bytes()
    source = {
        "path": str(resolved),
        "sha256": sha256_bytes(data),
        "size": len(data),
        "extension": resolved.suffix.lower(),
    }
    try:
        document = parse_document(data, resolved.name, options)
        status = status_for_document(document)
        warnings = list(document.pop("warnings", []))
    except (OSError, ParseError, UnicodeError) as exc:
        document = {
            "format": detect_format(data, resolved.name),
            "backend": "none",
            "metadata": {},
            "units": [],
            "members": [],
            "error": str(exc),
        }
        status = "FAILED"
        warnings = [str(exc)]
    unit_count = len(document.get("units", []))
    member_count = len(document.get("members", []))
    return {
        "schema_version": SCHEMA_VERSION,
        "parser": {
            "name": PARSER_NAME,
            "version": PARSER_VERSION,
            "parsed_at": utc_now(),
        },
        "source": source,
        "status": status,
        "document": document,
        "warnings": warnings,
        "stats": {"unit_count": unit_count, "member_count": member_count},
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Normalize a PDF, JSON, CSV, spreadsheet, text, image, or ZIP"
    )
    parser.add_argument("--input", type=Path)
    parser.add_argument("--output", type=Path)
    parser.add_argument(
        "--capabilities",
        action="store_true",
        help="Print available parser backends and exit",
    )
    parser.add_argument(
        "--password-env",
        help="Read a PDF/ZIP password from this environment variable",
    )
    parser.add_argument(
        "--ocr", choices=("auto", "always", "never"), default="auto"
    )
    parser.add_argument("--max-rows", type=int, default=Limits.max_rows)
    parser.add_argument(
        "--max-json-leaves", type=int, default=Limits.max_json_leaves
    )
    parser.add_argument(
        "--max-file-bytes", type=int, default=Limits.max_file_bytes
    )
    return parser


def parser_capabilities() -> dict[str, Any]:
    return {
        "parser_version": PARSER_VERSION,
        "formats": {
            "json_csv_text_xml_html_docx_zip": {
                "available": True,
                "backend": "python-standard-library",
            },
            "xlsx": {
                "available": True,
                "primary": bool(importlib.util.find_spec("openpyxl")),
                "fallback": "python-ooxml",
            },
            "xls": {
                "available": bool(importlib.util.find_spec("xlrd")),
                "backend": "xlrd",
            },
            "pdf": {
                "available": bool(importlib.util.find_spec("fitz"))
                or bool(shutil.which("pdftotext")),
                "pymupdf": bool(importlib.util.find_spec("fitz")),
                "pdftotext": bool(shutil.which("pdftotext")),
            },
            "image": {
                "metadata": bool(importlib.util.find_spec("PIL")),
                "ocr": bool(shutil.which("tesseract")),
            },
            "aes_zip": {
                "available": bool(importlib.util.find_spec("pyzipper")),
                "backend": "pyzipper",
            },
        },
    }


def main() -> int:
    argument_parser = build_parser()
    args = argument_parser.parse_args()
    if args.capabilities:
        print(json.dumps(parser_capabilities(), indent=2, sort_keys=True))
        return 0
    if args.input is None or args.output is None:
        argument_parser.error("--input and --output are required unless using --capabilities")
    password = os.environ.get(args.password_env) if args.password_env else None
    if args.password_env and password is None:
        print(
            f"error: environment variable {args.password_env!r} is not set",
            file=sys.stderr,
        )
        return 2
    options = Options(
        password=password,
        ocr=args.ocr,
        limits=Limits(
            max_file_bytes=args.max_file_bytes,
            max_rows=args.max_rows,
            max_json_leaves=args.max_json_leaves,
        ),
    )
    try:
        envelope = parse_path(args.input, options)
        atomic_json(args.output, envelope)
    except OSError as exc:
        print(f"error: {exc}", file=sys.stderr)
        return 2
    print(
        f"{envelope['status']}: {envelope['document']['format']} via "
        f"{envelope['document']['backend']} -> {args.output}"
    )
    return 0 if envelope["status"] != "FAILED" else 2


if __name__ == "__main__":
    raise SystemExit(main())
